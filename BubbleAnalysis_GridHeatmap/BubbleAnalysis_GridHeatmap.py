import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from tkinter import Tk, filedialog

def select_csv_file():
    """ユーザーがGUIでCSVファイルを選択する。"""
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        filetypes=[("CSV files", "*.csv")],
        title="CSVファイルを選択してください"
    )
    return file_path

def input_observation_area():
    """ユーザーから観察領域のY座標の最小値と最大値を入力し、X座標を決定する。"""
    print("観察領域のY座標の最小値と最大値を入力してください。")
    y_min = float(input("Y座標の最小値: "))
    y_max = float(input("Y座標の最大値: "))
    x_min = float(input("X座標の最小値: "))
    x_max = float(input("X座標の最大値: "))
    
    return x_min, x_max, y_min, y_max

def read_csv_data(csv_file):
    """CSVファイルを読み込み、ヘッダーを除いたデータを取得する。"""
    # ヘッダー行を特定して読み込む
    with open(csv_file, 'r', encoding='mbcs') as f:
        lines = f.readlines()
        for idx, line in enumerate(lines):
            if '時間,' in line:
                header_idx = idx
                break
    # データを読み込む
    df = pd.read_csv(csv_file, skiprows=header_idx, encoding='mbcs')
    return df

def preprocess_data(df):
    """データの前処理を行う。欠損値の処理や列名の整形を行う。"""
    # 列名を整える
    df.columns = df.columns.str.strip()
    df = df.replace('*', np.nan)  # '*'を欠損値に置換
    df = df.dropna(how='all')  # 全てがNaNの行を削除
    return df

def extract_bubble_data(df):
    """粒子数が変動しても対応できるように気泡データを抽出する。"""
    bubble_data = []
    # 'P'で始まる列を抽出
    position_columns = [col for col in df.columns if '(X)' in col]
    for col in position_columns:
        base_name = col[:-3]  # 'P1'などを取得
        x_col = f'{base_name}(X)'
        y_col = f'{base_name}(Y)'
        area_col = f'{base_name}(Area)'
        if x_col in df.columns and y_col in df.columns and area_col in df.columns:
            temp_df = df[[x_col, y_col, area_col]].copy()
            temp_df.columns = ['X', 'Y', 'Area']
            # 欠損値の処理とデータ型の変換
            temp_df = temp_df.replace('*', np.nan)
            temp_df['X'] = pd.to_numeric(temp_df['X'], errors='coerce')
            temp_df['Y'] = pd.to_numeric(temp_df['Y'], errors='coerce')
            temp_df['Area'] = pd.to_numeric(temp_df['Area'], errors='coerce')
            temp_df.dropna(inplace=True)
            bubble_data.append(temp_df)
    if not bubble_data:
        print("有効な気泡データが見つかりませんでした。")
        exit()
    bubble_df = pd.concat(bubble_data, ignore_index=True)
    return bubble_df

def calculate_diameter_and_volume(bubble_df):
    """面積から直径を計算し、体積を計算する。"""
    # 面積から直径を計算（円と仮定）
    bubble_df['Diameter'] = 2 * np.sqrt(bubble_df['Area'] / np.pi)
    
    # 体積を計算（計算式は空欄）
    def calculate_volume(diameter):
        if diameter > 100:
            volume = np.pi * ((diameter / 2) ** 2 ) * 100  # 底面がdiameterの円，高さが加工溝幅（100umとする）の円柱の体積
        else:
            volume = (4 * np.pi / 3) * ((diameter / 2) ** 3 )  # 球の体積
        return volume
    
    bubble_df['Volume'] = bubble_df['Diameter'].apply(calculate_volume)
    return bubble_df

def filter_bubbles_in_area(bubble_df, x_min, x_max, y_min, y_max):
    """観察領域内の気泡データに限定する。"""
    bubble_df = bubble_df[
        (bubble_df['X'] >= x_min) & (bubble_df['X'] <= x_max) &
        (bubble_df['Y'] >= y_min) & (bubble_df['Y'] <= y_max)
    ]
    return bubble_df

def aggregate_data(bubble_df, x_min, x_max, y_min, y_max):
    """データをグリッドごとに集計する。"""
    grid_size = 10
    x_bins = np.linspace(x_min, x_max, grid_size + 1)
    y_bins = np.linspace(y_min, y_max, grid_size + 1)
    
    # グリッド番号を割り当て
    bubble_df['X_bin'] = pd.cut(bubble_df['X'], bins=x_bins, labels=False, include_lowest=True)
    bubble_df['Y_bin'] = pd.cut(bubble_df['Y'], bins=y_bins, labels=False, include_lowest=True)
    
    # 集計
    total_grids = grid_size * grid_size
    processed_grids = 0
    results = []
    print("集計を開始します...")
    for x in range(grid_size):
        for y in range(grid_size):
            grid_data = bubble_df[
                (bubble_df['X_bin'] == x) & (bubble_df['Y_bin'] == y)
            ]
            bubble_count = grid_data['Diameter'].count()
            average_diameter = grid_data['Diameter'].mean()
            total_volume = grid_data['Volume'].sum()
            results.append({
                'X_bin': x,
                'Y_bin': y,
                'Bubble_Count': bubble_count,
                'Average_Diameter': average_diameter,
                'Total_Volume': total_volume
            })
            # 進行状況の表示
            processed_grids += 1
            progress = (processed_grids / total_grids) * 100
            print(f"進行状況: {progress:.2f}% ({processed_grids}/{total_grids})", end='\r')
    print("\n集計が完了しました。")
    grid_stats = pd.DataFrame(results)
    return grid_stats

def calculate_overall_statistics(bubble_df):
    """全体の統計情報を計算する。"""
    total_bubbles = bubble_df['Diameter'].count()
    average_diameter_overall = bubble_df['Diameter'].mean()
    total_volume_overall = bubble_df['Volume'].sum()
    diameter_std = bubble_df['Diameter'].std()
    diameter_min = bubble_df['Diameter'].min()
    diameter_max = bubble_df['Diameter'].max()
    return {
        'total_bubbles': total_bubbles,
        'average_diameter': average_diameter_overall,
        'total_volume': total_volume_overall,
        'diameter_std': diameter_std,
        'diameter_min': diameter_min,
        'diameter_max': diameter_max
    }

def save_to_excel(grid_stats, overall_stats, csv_file):
    """結果をExcelファイルに保存する。"""
    # ファイル名とパスの設定
    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    dir_name = os.path.dirname(csv_file)
    excel_file = os.path.join(dir_name, f"{base_name}_result.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "集計結果"
    
    # ヘッダーを書き込み
    ws.append(['X_bin', 'Y_bin', 'Bubble_Count', 'Average_Diameter', 'Total_Volume'])
    
    # データを書き込み
    for index, row in grid_stats.iterrows():
        ws.append([
            row['X_bin'], row['Y_bin'],
            row['Bubble_Count'],
            row['Average_Diameter'],
            row['Total_Volume']
        ])
    
    # サマリーシートの作成
    summary_ws = wb.create_sheet(title="サマリー")
    summary_ws.append(["全体の気泡数", overall_stats['total_bubbles']])
    summary_ws.append(["平均直径", overall_stats['average_diameter']])
    summary_ws.append(["直径の標準偏差", overall_stats['diameter_std']])
    summary_ws.append(["直径の最小値", overall_stats['diameter_min']])
    summary_ws.append(["直径の最大値", overall_stats['diameter_max']])
    summary_ws.append(["体積の合計", overall_stats['total_volume']])
    
    # ヒートマップシートの作成
    create_heatmap_sheet(wb, grid_stats)
    
    # Excelファイルの保存
    wb.save(excel_file)
    print(f"Excelファイルを保存しました: {excel_file}")

def create_heatmap_sheet(wb, grid_stats):
    """ヒートマップを作成する。"""
    heatmap_ws = wb.create_sheet(title="ヒートマップ")
    metrics = ['Bubble_Count', 'Average_Diameter', 'Total_Volume']
    grid_size = 10
    
    for idx, metric in enumerate(metrics):
        # 各メトリックごとに10x10のヒートマップを作成
        start_row = idx * (grid_size + 3) + 2  # 各ヒートマップの開始行を調整
        start_col = 2  # 開始列
        
        # タイトルを設定
        heatmap_ws.cell(row=start_row - 1, column=start_col, value=metric)
        
        # X_binとY_binのリスト（逆順）
        x_bins = list(range(grid_size - 1, -1, -1))  # 9から0へ
        y_bins = list(range(grid_size))  # 0から9へ
        
        # X_binのラベルを設定（上部に表示）
        for col_index, x_bin in enumerate(x_bins):
            heatmap_ws.cell(row=start_row, column=start_col + col_index + 1, value=str(x_bin))
            heatmap_ws.cell(row=start_row, column=start_col + col_index + 1).alignment = Alignment(horizontal='center')
        
        # Y_binのラベルを設定（左側に表示）
        for row_index, y_bin in enumerate(y_bins):
            heatmap_ws.cell(row=start_row + grid_size - row_index, column=start_col, value=str(y_bin))
            heatmap_ws.cell(row=start_row + grid_size - row_index, column=start_col).alignment = Alignment(horizontal='center')
        
        # ヒートマップのデータを埋め込む
        for _, row in grid_stats.iterrows():
            x_bin = int(row['X_bin'])
            y_bin = int(row['Y_bin'])
            value = row[metric]
            if pd.isna(value):
                value = 0  # 欠損値は0とする
            # Excel上のセル位置を計算
            col_pos = start_col + x_bins.index(x_bin) + 1  # +1はラベル列があるため
            row_pos = start_row + grid_size - y_bin  # Y_binは下から上へ
            cell = heatmap_ws.cell(row=row_pos, column=col_pos, value=value)
            # セルの値を非表示にする（フォント色を背景と同じにする）
            cell.number_format = ';;;'  # 表示形式を非表示に設定
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ヒートマップの範囲
        data_range = f"{heatmap_ws.cell(row=start_row + 1, column=start_col + 1).coordinate}:{heatmap_ws.cell(row=start_row + grid_size, column=start_col + grid_size).coordinate}"
        
        # 条件付き書式を設定（白黒、値が大きいほど白く）
        black_white_rule = ColorScaleRule(start_type='min', start_color='000000', end_type='max', end_color='FFFFFF')
        heatmap_ws.conditional_formatting.add(data_range, black_white_rule)
    
    # 列幅と行高さの調整
    for col in heatmap_ws.columns:
        heatmap_ws.column_dimensions[col[0].column_letter].width = 4
    for row in heatmap_ws.rows:
        heatmap_ws.row_dimensions[row[0].row].height = 20

def main():
    csv_file = select_csv_file()
    if not csv_file:
        print("CSVファイルが選択されませんでした。")
        exit()
    x_min, x_max, y_min, y_max = input_observation_area()
    df = read_csv_data(csv_file)
    df = preprocess_data(df)
    bubble_df = extract_bubble_data(df)
    bubble_df = calculate_diameter_and_volume(bubble_df)
    bubble_df = filter_bubbles_in_area(bubble_df, x_min, x_max, y_min, y_max)
    grid_stats = aggregate_data(bubble_df, x_min, x_max, y_min, y_max)
    overall_stats = calculate_overall_statistics(bubble_df)
    save_to_excel(grid_stats, overall_stats, csv_file)

if __name__ == "__main__":
    main()
