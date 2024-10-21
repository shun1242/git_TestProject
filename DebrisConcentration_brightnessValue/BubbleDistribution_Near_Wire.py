import os
import numpy as np
import pandas as pd
from PIL import Image
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

def select_bmp_files():
    root = Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(title="Select BMP Files", filetypes=[("BMP files", "*.bmp")])
    return root.tk.splitlist(file_paths)

def extract_brightness(image_path, y_coordinate, x_start_coordinate, x_end_coordinate):
    image = Image.open(image_path)
    image_array = np.array(image)
    return image_array[y_coordinate, x_start_coordinate:x_end_coordinate]

def rotate_90_clockwise(data):
    return np.rot90(data.reshape(1, -1), k=-1)

def main():
    # ユーザーに画像ファイルを選択させる
    bmp_files = select_bmp_files()
    if not bmp_files:
        print("No files selected.")
        return

    # ユーザーにy座標を入力させる
    y_coordinate = int(input("Enter the y-coordinate: "))

    # ユーザーにx座標の範囲を入力させる
    x_start_coordinate = int(input("Enter the x start coordinate: "))
    x_end_coordinate = int(input("Enter the x end coordinate: "))

    # 輝度値を格納するためのリスト
    brightness_data = None

    # 各画像から輝度値を抽出して回転、そしてリストに追加
    for bmp_file in bmp_files:
        brightness_row = extract_brightness(bmp_file, y_coordinate, x_start_coordinate, x_end_coordinate)
        rotated_data = rotate_90_clockwise(brightness_row)
        if brightness_data is None:
            brightness_data = rotated_data
        else:
            brightness_data = np.hstack((brightness_data, rotated_data))

    # 輝度値データをExcelファイルに出力
    df = pd.DataFrame(brightness_data)
    excel_path = os.path.join(os.path.dirname(bmp_files[0]), 'result.xlsx')
    df.to_excel(excel_path, index=False, header=False)

    # Excelファイルをオープンしてフォーマットを設定
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 色のスケールを設定
    max_row, max_col = brightness_data.shape
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='000000',
                                      end_type='num', end_value=255, end_color='FFFFFF')
    ws.conditional_formatting.add(f'A1:{end_cell}', color_scale_rule)

    # セルのサイズを設定
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 5  # 高さ5ピクセル
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 0.83  # 幅5ピクセル相当

    wb.save(excel_path)

    print(f"Brightness data has been saved to {excel_path}")

if __name__ == "__main__":
    main()