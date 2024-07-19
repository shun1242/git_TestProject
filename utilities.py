# utilities.py

import os
import cv2
import numpy as np
import pandas as pd
from tkinter import filedialog, Tk, simpledialog
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# BubbleDistribution_Near_Wire.py の関数

def select_bmp_files():
    """
    ユーザーにBMPファイルを選択させ、そのファイルパスのリストを返す。
    
    Returns:
    - list: 選択されたBMPファイルのパスのリスト
    """
    root = Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(title="Select BMP Files", filetypes=[("BMP files", "*.bmp")])
    return root.tk.splitlist(file_paths)

def extract_brightness(image_path, y_coordinate, x_start_coordinate, x_end_coordinate):
    """
    指定されたBMPファイルから特定のy座標およびx座標範囲の輝度値を抽出する。
    
    Args:
    - image_path (str): 画像ファイルのパス
    - y_coordinate (int): 抽出するy座標
    - x_start_coordinate (int): 抽出するx座標の開始位置
    - x_end_coordinate (int): 抽出するx座標の終了位置
    
    Returns:
    - numpy.ndarray: 抽出された輝度値の配列
    """
    image = Image.open(image_path)
    image_array = np.array(image)
    return image_array[y_coordinate, x_start_coordinate:x_end_coordinate]

def save_brightness_to_excel(brightness_data, output_path):
    """
    輝度データをExcelファイルに保存し、セルのフォーマットを設定する。
    
    Args:
    - brightness_data (numpy.ndarray): 輝度データの配列
    - output_path (str): 保存するExcelファイルのパス
    """
    df = pd.DataFrame(brightness_data)
    df.to_excel(output_path, index=False, header=False)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    max_row, max_col = brightness_data.shape
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='000000',
                                      end_type='num', end_value=255, end_color='FFFFFF')
    ws.conditional_formatting.add(f'A1:{end_cell}', color_scale_rule)

    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 5
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 0.83

    wb.save(output_path)

# calculate_AverageBrightness.py の関数

def select_folder():
    """
    ユーザーにフォルダを選択させ、そのパスを返す。
    
    Returns:
    - str: 選択されたフォルダのパス
    """
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory()
    return folder_selected

def load_images_from_folder(folder_selected):
    """
    指定されたフォルダから画像ファイルのリストを取得する。
    
    Args:
    - folder_selected (str): フォルダのパス
    
    Returns:
    - list: フォルダ内の画像ファイルのパスのリスト
    """
    return [os.path.join(folder_selected, f) for f in os.listdir(folder_selected) if f.endswith(('.png', '.jpg', '.jpeg', '.bmp'))]

def calculate_average_brightness(images):
    """
    複数の画像から平均輝度を計算し、平均化された画像を返す。
    
    Args:
    - images (list): 画像ファイルのパスのリスト
    
    Returns:
    - numpy.ndarray: 平均輝度を計算した画像の配列
    """
    sum_array = None
    count = 0
    
    for image_path in images:
        img = cv2.imread(image_path)
        if img is None:
            continue
        if img.ndim == 3:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        if sum_array is None:
            sum_array = np.zeros_like(img, dtype=np.float64)
        elif img.shape != sum_array.shape:
            continue

        sum_array += img
        count += 1

    if count == 0:
        return None

    average_array = sum_array / count
    return np.clip(average_array, 0, 255).astype(np.uint8)

# create_movie.py の関数

def create_movie_from_images(image_files, output_file, fps, frame_size):
    """
    画像ファイルのリストから動画を作成し、指定されたファイルに保存する。
    
    Args:
    - image_files (list): 画像ファイルのパスのリスト
    - output_file (str): 出力する動画ファイルのパス
    - fps (int): フレームレート
    - frame_size (tuple): 画像サイズ (幅, 高さ)
    """
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter(output_file, fourcc, fps, frame_size)

    for image_file in image_files:
        frame = cv2.imread(image_file)
        out.write(frame)

    out.release()
    cv2.destroyAllWindows()

# adjust_contrast.py の関数

def adjust_contrast(image, lower_bound=200, upper_bound=255, new_lower_bound=120, new_upper_bound=255):
    """
    画像のコントラストを調整し、新しい画像を返す。
    
    Args:
    - image (numpy.ndarray): 入力画像
    - lower_bound (int): 元の下限輝度値
    - upper_bound (int): 元の上限輝度値
    - new_lower_bound (int): 新しい下限輝度値
    - new_upper_bound (int): 新しい上限輝度値
    
    Returns:
    - numpy.ndarray: コントラストが調整された画像
    """
    grayscale_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    image_arr = np.array(grayscale_image)
    a = (image_arr - lower_bound) / (upper_bound - lower_bound)
    new_image_arr = a * (new_upper_bound - new_lower_bound) + new_lower_bound
    return np.clip(new_image_arr, 0, 255).astype(np.uint8)

# apply_csv_AverageFilter.py の関数

def select_file_and_apply_filter():
    """
    ユーザーにCSVファイルを選択させ、平均化フィルタを適用する。
    
    Returns:
    - tuple: フィルタが適用されたデータ (numpy.ndarray), ファイルパス (str), フィルタサイズ (int)
    """
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title='CSVファイルを選択してください', filetypes=[("CSV files", "*.csv")])
    if not file_path:
        return None, None

    filter_size = simpledialog.askinteger("フィルタサイズ入力", "平均化フィルタのサイズを入力してください (例: 3):", minvalue=1)
    if not filter_size:
        return None, None

    data = pd.read_csv(file_path)
    array_data = data.values
    filtered_data = uniform_filter(array_data, size=filter_size, mode='nearest')
    return filtered_data, file_path, filter_size

def save_filtered_data_to_csv(filtered_data, file_path, filter_size):
    """
    平均化フィルタが適用されたデータをCSVファイルに保存する。
    
    Args:
    - filtered_data (numpy.ndarray): フィルタが適用されたデータ
    - file_path (str): 元のCSVファイルのパス
    - filter_size (int): フィルタサイズ
    
    Returns:
    - str: 保存されたCSVファイルのパス
    """
    dirname = os.path.dirname(file_path)
    basename = os.path.basename(file_path)
    name_part = os.path.splitext(basename)[0]
    new_filename = f"{name_part}_AveragingFilter({filter_size}).csv"
    new_file_path = os.path.join(dirname, new_filename)
    pd.DataFrame(filtered_data).to_csv(new_file_path, index=False)
    return new_file_path

# DebrisConcentDistribution.py の関数

def extract_brightness_region(image_path, x1, y1, x2, y2):
    """
    指定された画像ファイルから特定の座標範囲の輝度値を抽出する。
    
    Args:
    - image_path (str): 画像ファイルのパス
    - x1, y1, x2, y2 (int): 抽出する範囲の座標
    
    Returns:
    - numpy.ndarray: 抽出された輝度値の配列
    """
    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"画像ファイル '{image_path}' の読み込みに失敗しました。")
    grayscale_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    return grayscale_image[y1:y2, x1:x2]

def calculate_debris_concentration(brightness_array):
    """
    輝度値から加工粉濃度を算出する。
    
    Args:
    - brightness_array (numpy.ndarray): 輝度値の配列
    
    Returns:
    - numpy.ndarray: 加工粉濃度の配列
    """
    return np.where(brightness_array > 222, 
                    10.997 - 4.5687 * np.log10(brightness_array), 
                    77.496 - 32.887 * np.log10(brightness_array))

# save_frame_range.py の関数

def save_frame_range(video_path, start_frame, stop_frame, step_frame, dir_path, basename):
    """
    指定された範囲のフレームをビデオファイルから抽出し、グレースケールで画像として保存する。
    
    Args:
    - video_path (str): ビデオファイルのパス
    - start_frame (int): 抽出を開始するフレーム番号
    - stop_frame (int): 抽出を終了するフレーム番号
    - step_frame (int): 抽出するフレームの間隔
    - dir_path (str): フレームを保存するディレクトリのパス
    - basename (str): 保存するファイルのベース名
    """
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"ビデオファイル '{video_path}' を開けませんでした。")

    os.makedirs(dir_path, exist_ok=True)
    base_path = os.path.join(dir_path, basename)

    for n in range(start_frame, stop_frame, step_frame):
        cap.set(cv2.CAP_PROP_POS_FRAMES, n)
        ret, frame = cap.read()
        if ret:
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            file_path = f'{base_path}_{n}.bmp'
            cv2.imwrite(file_path, gray_frame)
        else:
            print('フレームの読み込みに失敗しました。')
            break

    cap.release()
    cv2.destroyAllWindows()
    print('処理が完了しました。')